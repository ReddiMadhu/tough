"""
Excel Migration Report Generator — 5 sheets:
  1. DAX Conversions
  2. Data Model
  3. Visualizations (informational — visuals not migrated)
  4. Migration Notes
  5. Formula Dependency Graph
"""
import json
from pathlib import Path
from typing import List, Dict, Any
from loguru import logger

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed — Excel report will be skipped")


class ExcelReportGenerator:
    """Generate a 5-sheet Excel migration report."""

    # Color palette
    HEADER_FILL = "1E40AF"     # indigo-800
    GREEN_FILL = "DCFCE7"      # green-100
    YELLOW_FILL = "FEF9C3"    # yellow-100
    RED_FILL = "FEE2E2"        # red-100
    ALT_ROW_FILL = "F8FAFC"   # slate-50

    def generate(
        self,
        intermediate_model: Dict[str, Any],
        dax_conversions: List[Dict[str, Any]],
        output_dir: str,
        migration_id: str,
    ) -> str:
        """Generate the Excel report and return the file path."""
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel report generation")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        self._sheet_business_narrative(wb, intermediate_model)
        self._sheet_dax_conversions(wb, dax_conversions)
        self._sheet_data_model(wb, intermediate_model)
        self._sheet_visualizations(wb, intermediate_model)
        self._sheet_migration_notes(wb, dax_conversions, intermediate_model)
        self._sheet_formula_dependency(wb, dax_conversions)

        output_path = Path(output_dir) / f"{migration_id}_migration_report.xlsx"
        wb.save(str(output_path))
        logger.info(f"Excel report saved: {output_path}")
        return str(output_path)


    # ── Sheet 1: DAX Conversions ───────────────────────────────────────────────

    def _sheet_dax_conversions(self, wb: "Workbook", conversions: List[Dict]):
        ws = wb.create_sheet("DAX Conversions")
        headers = [
            "Measure Name", "ThoughtSpot Formula", "DAX Formula",
            "Confidence %", "Pattern", "Requires Review",
            "Source Object", "Source Object Type", "Notes",
        ]
        self._write_header_row(ws, headers)

        for i, conv in enumerate(conversions, start=2):
            confidence = conv.get("confidence", 0)
            requires_review = conv.get("requires_review", False)
            notes = conv.get("notes", [])
            notes_str = "; ".join(notes) if isinstance(notes, list) else str(notes)

            row = [
                conv.get("measure_name", ""),
                conv.get("original_formula", ""),
                conv.get("dax_formula", ""),
                f"{confidence:.0%}",
                conv.get("pattern", ""),
                "Yes" if requires_review else "No",
                conv.get("source_object", ""),
                conv.get("source_object_type", ""),
                notes_str,
            ]
            ws.append(row)

            # Traffic-light confidence coloring
            fill_color = (
                self.GREEN_FILL if confidence >= 0.9 else
                self.YELLOW_FILL if confidence >= 0.6 else
                self.RED_FILL
            )
            fill = PatternFill("solid", fgColor=fill_color)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=i, column=col_idx).fill = fill
                ws.cell(row=i, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

        self._autofit_columns(ws, [30, 50, 60, 14, 22, 16, 30, 20, 60])

    # ── Sheet 2: Data Model ────────────────────────────────────────────────────

    def _sheet_data_model(self, wb: "Workbook", model: Dict):
        ws = wb.create_sheet("Data Model")
        headers = [
            "Table Name", "Column Name", "Data Type (ThoughtSpot)",
            "Power BI Type", "Column Role", "Source Column",
            "Aggregation", "Format Pattern",
        ]
        self._write_header_row(ws, headers)

        DT_MAP = {
            "VARCHAR": "string", "CHAR": "string", "TEXT": "string",
            "INT32": "int64", "INT64": "int64", "INTEGER": "int64",
            "DOUBLE": "double", "FLOAT": "double", "DECIMAL": "double",
            "BOOL": "boolean", "DATE": "dateTime", "DATETIME": "dateTime",
            "TIMESTAMP": "dateTime", "TIME": "dateTime",
        }

        row_num = 2
        for table in model.get("tables", []):
            table_name = table.get("name", "")
            for col in table.get("column_details", []):
                ts_type = col.get("data_type", "VARCHAR")
                pbi_type = DT_MAP.get(ts_type.upper() if ts_type else "", "string")
                ws.append([
                    table_name,
                    col.get("name", ""),
                    ts_type,
                    pbi_type,
                    col.get("column_type", "ATTRIBUTE"),
                    col.get("db_column_name", col.get("name", "")),
                    col.get("aggregation", ""),
                    col.get("format_pattern", ""),
                ])
                if row_num % 2 == 0:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=c).fill = PatternFill("solid", fgColor=self.ALT_ROW_FILL)
                row_num += 1

        # Relationships section
        joins = model.get("joins", [])
        if joins:
            ws.append([])
            ws.append(["--- RELATIONSHIPS ---"])
            ws.append(["Join Name", "Left Table", "Left Column", "Right Table", "Right Column", "Join Type", "Cardinality"])
            for join in joins:
                ws.append([
                    join.get("name", ""),
                    join.get("left_table", ""),
                    join.get("left_column", ""),
                    join.get("right_table", ""),
                    join.get("right_column", ""),
                    join.get("join_type", ""),
                    join.get("cardinality", ""),
                ])

        self._autofit_columns(ws, [25, 30, 25, 15, 20, 30, 15, 20])

    # ── Sheet 3: Visualizations ────────────────────────────────────────────────

    def _sheet_visualizations(self, wb: "Workbook", model: Dict):
        ws = wb.create_sheet("Visualizations")
        ws.append(["NOTE: Visuals are NOT migrated. This sheet is informational only."])
        ws.cell(row=1, column=1).font = Font(bold=True, color="DC2626")
        ws.append([])

        headers = [
            "Source Liveboard / Answer", "Visualization Name",
            "TS Chart Type", "Migration Status", "Search Query",
        ]
        self._write_header_row(ws, headers)

        for ws_entry in model.get("worksheets", []):
            ws.append([
                ws_entry.get("source_liveboard", ws_entry.get("source_object_type", "")),
                ws_entry.get("name", ""),
                ws_entry.get("ts_chart_type", ""),
                "Not migrated — add manually in Power BI Desktop",
                ws_entry.get("search_query", ""),
            ])

        self._autofit_columns(ws, [35, 35, 20, 45, 60])

    # ── Sheet 4: Migration Notes ───────────────────────────────────────────────

    def _sheet_migration_notes(self, wb: "Workbook", conversions: List[Dict], model: Dict):
        ws = wb.create_sheet("Migration Notes")
        headers = ["Category", "Object", "Message", "Recommendation"]
        self._write_header_row(ws, headers)

        notes = []

        # Formula review required
        for conv in conversions:
            if conv.get("requires_review"):
                notes.append([
                    "Action Required",
                    conv.get("measure_name", ""),
                    f"Formula requires manual DAX authoring (pattern: {conv.get('pattern', '')})",
                    "Open Power BI Desktop and edit the measure in the DAX editor",
                ])

        # query_groups() formulas
        for conv in conversions:
            if conv.get("pattern") == "DYNAMIC_GROUPING":
                notes.append([
                    "Warning",
                    conv.get("measure_name", ""),
                    "Formula uses query_groups() — no direct DAX equivalent",
                    "Manually author using CALCULATE with explicit ALLEXCEPT or REMOVEFILTERS",
                ])

        # Visuals not migrated
        viz_count = len(model.get("worksheets", []))
        if viz_count > 0:
            notes.append([
                "Info",
                "Report Visuals",
                f"{viz_count} visualizations found in ThoughtSpot but NOT migrated (out of scope)",
                "Add visuals manually in Power BI Desktop using the Measures table",
            ])

        # Connection info ignored
        notes.append([
            "Info",
            "Data Source Connection",
            "ThoughtSpot database connection info was not migrated",
            "Reconnect to your data source in Power BI Desktop (Get Data)",
        ])

        for note in notes:
            ws.append(note)

        self._autofit_columns(ws, [20, 30, 60, 60])

    # ── Sheet 5: Formula Dependency Graph ─────────────────────────────────────

    def _sheet_formula_dependency(self, wb: "Workbook", conversions: List[Dict]):
        ws = wb.create_sheet("Formula Dependency Graph")
        headers = ["Measure Name", "References Measures", "Referenced By", "Depth"]
        self._write_header_row(ws, headers)

        measure_names = {c.get("measure_name", "") for c in conversions}

        # Build reference map
        references: Dict[str, List[str]] = {}
        referenced_by: Dict[str, List[str]] = {}

        for conv in conversions:
            name = conv.get("measure_name", "")
            dax = conv.get("dax_formula", "")
            refs = []
            for other_name in measure_names:
                if other_name != name and other_name and f"[{other_name}]" in dax:
                    refs.append(other_name)
                    referenced_by.setdefault(other_name, []).append(name)
            references[name] = refs

        # Compute depth (BFS from leaf nodes)
        depth_map: Dict[str, int] = {}
        for name in measure_names:
            if not references.get(name):
                depth_map[name] = 0

        for conv in conversions:
            name = conv.get("measure_name", "")
            if name not in depth_map:
                depth_map[name] = max(
                    (depth_map.get(ref, 0) + 1 for ref in references.get(name, [])),
                    default=1,
                )

        for conv in conversions:
            name = conv.get("measure_name", "")
            refs = references.get(name, [])
            by = referenced_by.get(name, [])
            ws.append([
                name,
                ", ".join(refs) if refs else "(none)",
                ", ".join(by) if by else "(none)",
                depth_map.get(name, 0),
            ])

        self._autofit_columns(ws, [35, 50, 50, 10])

    # ── Sheet 6: Business Narrative ───────────────────────────────────────────

    def _sheet_business_narrative(self, wb: "Workbook", model: Dict):
        narrative = model.get("narrative_summary")
        if not narrative:
            return

        ws = wb.create_sheet("AI Business Narrative")
        ws.views.sheetView[0].showGridLines = True

        # Styled Title Block
        ws.cell(row=1, column=1, value="ThoughtSpot to Power BI — AI Executive Narrative").font = Font(size=16, bold=True, color="1E40AF")
        ws.row_dimensions[1].height = 30
        
        ws.cell(row=3, column=1, value="Executive Insight Report").font = Font(size=12, bold=True, color="475569")
        ws.cell(row=3, column=1).fill = PatternFill("solid", fgColor="F1F5F9")
        ws.row_dimensions[3].height = 20
        
        row_idx = 5
        lines = narrative.split("\n")
        for line in lines:
            line_str = line.strip()
            if not line_str:
                row_idx += 1
                continue
                
            cell = ws.cell(row=row_idx, column=1, value=line_str)
            if line_str.startswith("### "):
                cell.value = line_str.replace("### ", "")
                cell.font = Font(size=11, bold=True, color="1E40AF")
            elif line_str.startswith("## "):
                cell.value = line_str.replace("## ", "")
                cell.font = Font(size=12, bold=True, color="1E3A8A")
            elif line_str.startswith("# "):
                cell.value = line_str.replace("# ", "")
                cell.font = Font(size=13, bold=True, color="111827")
            elif line_str.startswith("- ") or line_str.startswith("* "):
                cell.font = Font(size=10, color="334155")
                cell.alignment = Alignment(indent=1)
            else:
                cell.font = Font(size=10, color="334155")
            row_idx += 1

        self._autofit_columns(ws, [100])

    # ── Helpers ────────────────────────────────────────────────────────────────


    def _write_header_row(self, ws, headers: List[str]):
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor=self.HEADER_FILL)
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _autofit_columns(self, ws, widths: List[int]):
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
